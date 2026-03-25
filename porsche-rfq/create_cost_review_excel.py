#!/usr/bin/env python3
"""Generate Cost Model Review Excel Report"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "/Users/maxshow/.openclaw/workspace/porsche-rfq/Cost_Model_Review_Report.xlsx"

# Colors
C_RED    = "C50000"
C_DARK   = "1A1A1A"
C_BLUE   = "005BB3"
C_F8     = "F8F8F8"
C_F0     = "F0F0F0"
C_WY     = "FFF2CC"   # warn yellow
C_WO     = "FCE4D6"   # warn orange
C_OK     = "E2EFDA"   # green
C_BORDER = "CCCCCC"

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, size=10, color=C_DARK, name="Calibri"):
    return Font(bold=bold, size=size, color=color, name=name)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def tb():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)

def thk():
    t = Side(style="thin", color=C_BORDER)
    b = Side(style="medium", color=C_DARK)
    return Border(left=t, right=t, top=t, bottom=b)

def hdr(ws, r, c, val, merge=None, bg=C_RED, fg="FFFFFF", sz=10):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(bg)
    cell.font = Font(bold=True, size=sz, color=fg, name="Calibri")
    cell.alignment = aln("center", "center")
    cell.border = tb()
    if merge:
        ws.merge_cells(start_row=r, start_column=c, end_row=r, end_column=merge)
    return cell

def sec(ws, r, c, val, merge=None):
    return hdr(ws, r, c, val, merge=merge, bg=C_DARK, fg="FFFFFF", sz=10)

def dat(ws, r, c, val, bg=C_F8, bold=False, color=C_DARK,
         ah="left", fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = fill(bg)
    cell.font = Font(bold=bold, size=10, color=color, name="Calibri")
    cell.alignment = aln(ah, "center")
    cell.border = tb()
    if fmt:
        cell.number_format = fmt
    return cell

def cw(ws, d):
    for col, w in d.items():
        ws.column_dimensions[get_column_letter(col)].width = w

wb = openpyxl.Workbook()

# ── Sheet 1: Executive Summary ─────────────────────────────────────────────
ws = wb.active
ws.title = "Executive Summary"
ws.sheet_view.showGridLines = False

for c in range(1,9): ws.cell(row=1,column=c).fill = fill(C_DARK)
ws.merge_cells("A1:H1")
t = ws.cell(row=1,column=1,value="保时捷上海临港 PDC 仓库项目 — 成本模型专业审查报告")
t.font = Font(bold=True,size=14,color="FFFFFF",name="Calibri")
t.alignment = aln("center","center")
ws.row_dimensions[1].height = 30

ws.merge_cells("A2:H2")
s = ws.cell(row=2,column=1,value="Porsche China Shanghai Lingang PDC — Cost Model Professional Review  |  V1.5 Max Update  |  2026-03-23")
s.fill = fill(C_RED); s.font = Font(size=10,color="FFFFFF",name="Calibri")
s.alignment = aln("center","center")

sec(ws,4,1,"核心指标摘要  |  KEY METRICS SUMMARY",merge=8)
hdrs = ["指标 / Metric","Y1","Y2","Y3","Y4","Y5","5年合计","审查意见"]
for c,h in enumerate(hdrs,1): hdr(ws,5,c,h)
ws.row_dimensions[5].height = 20

rows = [
    ("收入 / Revenue (RMB)","7,983,695","7,191,484","6,970,742","7,170,791","7,377,299","36,694,011","⚠️ 逐年小幅下降"),
    ("总成本 / Total Costs","7,858,293","6,603,872","6,595,080","6,555,944","6,758,295","34,371,484",""),
    ("EBITA","125,403","587,612","375,662","614,846","619,004","2,322,527","🔴 Y1极低"),
    ("EBITA率 / EBITA %","1.57%","8.17%","5.39%","8.57%","8.39%","6.33%","🔴 风险缓冲不足"),
    ("人工占比 / Labour %","60.2%","74.1%","74.2%","76.2%","77.0%","—","⚠️ 占比偏高"),
    ("KPI Malus敞口 / Malus Exposure","-234,000","-215,744","-209,122","-215,124","-221,332","—","🔴 可清零利润"),
]

for i,row in enumerate(rows):
    r = 6+i
    bg = C_F8 if i%2==0 else "FFFFFF"
    if "EBITA" in row[0] and "%" not in row[0]: bg = C_WO
    for c,val in enumerate(row,1):
        bold = (c==1)
        fc = C_DARK
        if c==8 and "🔴" in str(val): fc=C_RED; bold=True
        dat(ws,r,c,val,bg=bg,bold=bold,color=fc,ah="center" if c>1 else "left")
    ws.row_dimensions[r].height = 18

r=12
ws.merge_cells(f"A{r}:H{r}")
c=ws.cell(row=r,column=1,value="总体结论  |  OVERALL CONCLUSION")
c.fill=fill(C_DARK); c.font=Font(bold=True,size=11,color="FFFFFF",name="Calibri")
c.alignment=aln("left","center")

r=13
ws.merge_cells(f"A{r}:H{r}")
cell=ws.cell(row=r,column=1,value="该成本模型数据逻辑整体自洽，但 Y1 EBITA 率仅 1.57%，KPI Malus -3% 即可将全年利润清零。 建议优先处理：① DG入库/出库处理定价（P0）；② 责任险预算调整（P0）；③ 采纳减少1名Supervisor建议。综合调价+优化可将 Y1 EBITA 率提升至 8-10% 安全区间。")
cell.fill=fill(C_WY); cell.font=Font(size=10,color=C_DARK,name="Calibri")
cell.alignment=Alignment(horizontal="left",vertical="center",wrap_text=True)
cell.border=thk()
ws.row_dimensions[r].height = 55

cw(ws,{1:30,2:14,3:14,4:14,5:14,6:14,7:14,8:22})

# ── Sheet 2: Priority Actions ─────────────────────────────────────────────
ws2 = wb.create_sheet("Priority Actions")
ws2.sheet_view.showGridLines = False
for c in range(1,8): ws2.cell(row=1,column=c).fill=fill(C_DARK)
ws2.merge_cells("A1:G1")
t2=ws2.cell(row=1,column=1,value="优先处理事项  |  PRIORITY ACTION PLAN (P0-P2)")
t2.font=Font(bold=True,size=13,color="FFFFFF",name="Calibri"); t2.alignment=aln("center","center")
ws2.row_dimensions[1].height=28

ws2.merge_cells("A2:G2")
l=ws2.cell(row=2,column=1,value="🔴 P0=立即处理  |  🟡 P1=本周处理  |  🟢 P2=签约前处理")
l.fill=fill(C_F0); l.font=Font(size=10,color=C_DARK,name="Calibri"); l.alignment=aln("center","center")

hdrs2=["优先级","问题","当前值","建议值","增收/节支(RMB/年)","EBITA率提升","操作"]
for c,h in enumerate(hdrs2,1): hdr(ws2,3,c,h)
ws2.row_dimensions[3].height=20

p0 = [
    ("🔴 P0","DG入库处理单价 = 0（严重漏项）","0 RMB/行","10-12 RMB/行","+132K~+158K","+1.7%~+2.0%","Rate Sheet 新增项目"),
    ("🔴 P0","DG出库处理单价 = 0（P&L有收入383K）","0 RMB/行","12-15 RMB/行","+171K~+214K","+2.2%~+2.7%","Rate Sheet 新增项目"),
    ("🔴 P0","责任险预算严重低估（要求4亿覆盖）","80,000/年","350,000/年","成本+270K（风险消除）","-3.4%（消除风险）","MHE&INV List 更新保险预算"),
    ("🟡 P1","出库操作单价偏低","6.54 RMB/行","6.90 RMB/行","+252K","+3.2%","Rate Sheet 调整"),
    ("🟡 P1","减少1名 Supervisor（Parameter已标注）","2 Supervisor","1 Supervisor","+313K","+3.9%","Parameter Sheet 更新"),
    ("🟡 P1","中文贴标单价偏低","0.15 RMB/标签","0.22 RMB/标签","+138K","+1.7%","Rate Sheet 调整"),
    ("🟡 P1","Packing productivity 过激","39 line/hr","36 line/hr（保守）","降低不合规风险","间接改善","Prod Check 更新"),
    ("🟢 P2","DG Warranty Return 溢价","26.81 RMB/行","32 RMB/行","+4K","+0.05%","Rate Sheet 调整"),
    ("🟢 P2","Labour inflation Y4跳跃+9.4%","缺乏说明","平滑或加注","—","—","Parameter Sheet 添加注释"),
]
pc = {"🔴 P0":C_WO,"🟡 P1":C_WY,"🟢 P2":C_OK}
for i,row in enumerate(p0):
    r=4+i
    bg=pc.get(row[0],C_F8)
    for c,val in enumerate(row,1):
        bold=(c==1)
        fc=C_DARK
        if c==1 and "🔴" in val: fc=C_RED; bold=True
        dat(ws2,r,c,val,bg=bg,bold=bold,color=fc,ah="center" if c in (1,5,6,7) else "left")
    ws2.row_dimensions[r].height=20

r=14
ws2.merge_cells(f"A{r}:G{r}")
sm=ws2.cell(row=r,column=1,value="综合调价 + 成本优化 = Y1 EBITA 率从 1.57% 提升至 8-10%  |  Total Adjustments = Y1 EBITA from 1.57% → 8-10%")
sm.fill=fill(C_OK); sm.font=Font(bold=True,size=11,color=C_BLUE,name="Calibri")
sm.alignment=aln("center","center"); sm.border=thk()
ws2.row_dimensions[r].height=24
cw(ws2,{1:10,2:30,3:16,4:16,5:20,6:16,7:22})

# ── Sheet 3: EBITA Analysis ────────────────────────────────────────────────
ws3 = wb.create_sheet("EBITA Analysis")
ws3.sheet_view.showGridLines = False
for c in range(1,9): ws3.cell(row=1,column=c).fill=fill(C_DARK)
ws3.merge_cells("A1:H1")
t3=ws3.cell(row=1,column=1,value="EBITA 五年趋势与敏感性分析  |  EBITA 5-Year Trend & Sensitivity Analysis")
t3.font=Font(bold=True,size=13,color="FFFFFF",name="Calibri"); t3.alignment=aln("center","center")
ws3.row_dimensions[1].height=28

sec(ws3,3,1,"5年 P&L 汇总  |  5-Year P&L Summary",merge=8)
hdrs3=["项目 / Item","Y1","Y2","Y3","Y4","Y5","5年合计 / Total","占比 / Share"]
for c,h in enumerate(hdrs3,1): hdr(ws3,4,c,h)

pl=[
    ("收入 / Revenue",7983695,7191484,6970742,7170791,7377299,36694011,"100.0%"),
    ("  入库操作",1687469,1636845,1587740,1635372,1684433,8231860,"22.4%"),
    ("  出库操作",4575344,4438084,4304941,4434089,4567112,22319570,"60.8%"),
    ("  出库DG操作",382999,371509,360364,371174,382310,1868355,"5.1%"),
    ("  其他（报关/标签/WR等）",1340883,745046,717697,730156,743444,4258226,"11.6%"),
    ("总成本 / Total Costs",7858293,6603872,6595080,6555944,6758295,34371484,"100.0%"),
    ("  人工 / Labour",4729186,4897288,4894763,4994783,5202366,24718384,"71.9%"),
    ("    其中：Direct Labour",3319786,3445606,3399530,3454693,3616073,17235689,"50.1%"),
    ("    其中：Indirect Labour",1409400,1451682,1495232,1540089,1586292,7482695,"21.8%"),
    ("  设备 / Equipment",2741746,1706584,1700317,1561162,1555930,9265739,"27.0%"),
    ("    其中：折旧 / Depreciation",1163963,134958,133924,0,0,1432845,"4.2%"),
    ("  Lease/Consume",1577783,1571626,1566394,1561162,1555930,7832894,"22.8%"),
    ("  Startup",387360,0,0,0,0,387360,"1.1%"),
    ("EBITA",125403,587612,375662,614846,619004,2322527,"—"),
]

for i,row in enumerate(pl):
    r=5+i
    is_ebita = row[0]=="EBITA"
    is_sub = row[0].startswith("  ")
    is_rev = row[0] in ("收入 / Revenue","总成本 / Total Costs")
    bg = C_F8 if i%2==0 else "FFFFFF"
    if is_ebita: bg=C_WO
    elif is_rev: bg=C_F0
    for c,val in enumerate(row,1):
        bold=(c==1) or is_ebita
        fc=C_DARK
        if is_ebita and c==1: fc=C_RED; bold=True
        fmt='#,##0' if isinstance(val,(int,float)) and 1<c<8 else None
        dat(ws3,r,c,val,bg=bg,bold=bold,color=fc,ah="right" if (c>1 and c<8) else "left",fmt=fmt)
    ws3.row_dimensions[r].height=17

# Sensitivity
rs=20
sec(ws3,rs,1,"敏感性分析（Y1 EBITA 基准 = 125,403）|  SENSITIVITY ANALYSIS",merge=8)
hdrs4=["变量 / Variable","变动幅度 / Change","EBITA变化 / Δ","调整后EBITA","EBITA率","风险"]
for c,h in enumerate(hdrs4,1): hdr(ws3,rs+1,c,h)

sens=[
    ("基准 / Base","—","—","125,403","1.57%","🟡"),
    ("减少1名 Supervisor","-313K cost","+313K","438,403","5.5%","🟢"),
    ("KPI Malus -3%","-234K profit","-234K","-108,603","-1.4%","🔴"),
    ("货量 -5%","-400K revenue","-400K","-274,597","-3.4%","🔴"),
    ("责任险修正至350K","+270K cost","-270K","-144,597","-1.8%","🔴"),
    ("DG定价+10 RMB/行","+132K revenue","+132K","257,403","3.2%","🟡"),
    ("出库+0.36 RMB/行","+252K revenue","+252K","377,403","4.7%","🟡"),
    ("综合调价（P0+P1）","+700K revenue","+700K","825,403","10.3%","🟢"),
]
for i,row in enumerate(sens):
    r=rs+2+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    if "🔴" in row[5]: bg=C_WO
    elif "🟢" in row[5] and "基准" not in row[0]: bg=C_OK
    for c,val in enumerate(row,1):
        bold=(c==1)
        fc=C_DARK
        if "🔴" in str(val): fc=C_RED
        dat(ws3,r,c,val,bg=bg,bold=bold,color=fc,ah="right" if c in (2,3,4,5) else "left")
    ws3.row_dimensions[r].height=18

cw(ws3,{1:28,2:18,3:16,4:16,5:12,6:12,7:16,8:16})

# ── Sheet 4: Risk Matrix ────────────────────────────────────────────────
ws4 = wb.create_sheet("Risk Matrix")
ws4.sheet_view.showGridLines = False
for c in range(1,8): ws4.cell(row=1,column=c).fill=fill(C_DARK)
ws4.merge_cells("A1:G1")
t4=ws4.cell(row=1,column=1,value="风险矩阵  |  RISK MATRIX")
t4.font=Font(bold=True,size=13,color="FFFFFF",name="Calibri"); t4.alignment=aln("center","center")
ws4.row_dimensions[1].height=28

sec(ws4,3,1,"风险识别与量化  |  RISK IDENTIFICATION & QUANTIFICATION",merge=7)
hdrs5=["风险类别","风险项","当前估算","调整后估算","年度敞口(RMB)","严重度","应对措施"]
for c,h in enumerate(hdrs5,1): hdr(ws4,4,c,h)

risks=[
    ("定价漏项","DG入库处理 = 0","0/年","120K-172K/年","+120K~172K","🔴 高","Rate Sheet 新增DG入库单价"),
    ("定价漏项","DG出库处理 = 0（P&L有收入）","0/年","171K-214K/年","+171K~214K","🔴 高","Rate Sheet 新增DG出库单价"),
    ("成本低估","责任险（要求4亿覆盖）","80K/年","350K/年","+270K","🔴 高","保险预算修正+询价"),
    ("成本低估","人工流失/overtime风险","0","100K-150K/年","+100K~150K","🔴 高","薪资竞争力提升"),
    ("合同风险","KPI Malus -3%（关键KPI未达标）","0","-234K/年","-234K","🔴 高","KPI预警机制+学习期谈判"),
    ("数据逻辑","Outbound DG Rate=0但P&L有383K收入","逻辑矛盾","需统一","—","🟡 中","明确Outbound DG费率"),
    ("假设过激","Packing productivity 39（现34.3）","需提升13.7%","保守下调至36","降低不合规风险","🟡 中","Prod Check回调"),
    ("假设过激","Labour inflation Y4跳跃+9.4%","缺乏说明","平滑处理","—","🟡 中","Parameter Sheet添加注释"),
    ("遗漏成本","DG废弃物处理（普通42K统一）","含在通用预算","需单独估算","待定","🟡 中","向处置商询价"),
    ("遗漏成本","WMS系统许可证费用","含在IT running？","需确认","待定","🟢 低","向PCN确认系统归属"),
    ("运营风险","Y1磨合期FTE效率损失","已含在模型","—","—","🟡 中","预发人员提前到位"),
    ("市场风险","货量下滑（3+2年绑定）","年降3%","低货量保障条款","—","🟡 中","合同增设最低货量保护"),
]
for i,row in enumerate(risks):
    r=5+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    if "🔴" in row[5]: bg=C_WO
    elif "🟡" in row[5]: bg=C_WY
    for c,val in enumerate(row,1):
        bold=(c==1)
        fc=C_DARK
        if "🔴" in str(val): fc=C_RED; bold=True
        dat(ws4,r,c,val,bg=bg,bold=bold,color=fc)
    ws4.row_dimensions[r].height=22
cw(ws4,{1:14,2:28,3:16,4:16,5:16,6:10,7:30})

# ── Sheet 5: KPI Malus ─────────────────────────────────────────────────
ws5 = wb.create_sheet("KPI Malus Analysis")
ws5.sheet_view.showGridLines = False
for c in range(1,7): ws5.cell(row=1,column=c).fill=fill(C_DARK)
ws5.merge_cells("A1:F1")
t5=ws5.cell(row=1,column=1,value="KPI Bonus/Malus 风险分析  |  KPI BONUS/MALUS RISK ANALYSIS")
t5.font=Font(bold=True,size=13,color="FFFFFF",name="Calibri"); t5.alignment=aln("center","center")
ws5.row_dimensions[1].height=28

sec(ws5,3,1,"合同 Bonus/Malus 机制  |  CONTRACT BONUS/MALUS MECHANISM",merge=6)
hdrs6=["机制","触发条件","月度影响(Y1)","年度影响(Y1)","对EBITA影响","说明"]
for c,h in enumerate(hdrs6,1): hdr(ws5,4,c,h)

malus=[
    ("Bonus +3%","所有KPI达标","+19,500/月","+234,000/年","EBITA翻倍至359K",""),
    ("Malus -1%（非关键KPI）","1项+非关键KPI未达标","-6,500/月","-78,000/年","EBITA降至47K",""),
    ("Malus -3%（含关键KPI）","1项+关键KPI未达标","-19,500/月","-234,000/年","EBITA归零并亏损","🔴 Y1无缓冲"),
    ("Malus -3% ×2项","2项+关键KPI未达标","-39,000/月","-468,000/年","大幅亏损","🔴 灾难性"),
]
for i,row in enumerate(malus):
    r=5+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    if "亏损" in str(row[4]): bg=C_WO
    elif "翻倍" in str(row[4]): bg=C_OK
    for c,val in enumerate(row,1):
        bold=(c==1)
        fc=C_DARK
        if "🔴" in str(val): fc=C_RED
        dat(ws5,r,c,val,bg=bg,bold=bold,color=fc)
    ws5.row_dimensions[r].height=20

sec(ws5,10,1,"关键KPI未达标风险评估  |  KEY KPI FAILURE RISK",merge=6)
hdrs7=["KPI","PCN目标","飞力达承诺","未达标概率(Y1)","年度期望损失","风险等级"]
for c,h in enumerate(hdrs7,1): hdr(ws5,11,c,h)

kpis=[
    ("配送准确率 VOR","99.97%","99.98%","15-20%","35K-47K","🔴 高"),
    ("月度库存准确率","≤0.02%","≤0.01%","10-15%","23K-35K","🟡 中"),
    ("年度库存准确率","≤0.02%","≤0.01%","15-20%","35K-47K","🔴 高"),
    ("中转运输 KPI","海运3天/空运1天","同PCN","15-25%","35K-59K","🔴 高"),
    ("仓库索赔率","≤0.02%","0%","20-30%","47K-70K","🔴 高"),
    ("VOR出库时效","<25分钟","<25分钟","10-15%","—","🟡 中"),
    ("6S评分","≥95分","≥96分","5-10%","—","🟢 低"),
    ("合计期望损失","—","—","—","140K-235K","🔴 高"),
]
for i,row in enumerate(kpis):
    r=12+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    if "🔴" in row[5]: bg=C_WO
    elif "🟢" in row[5]: bg=C_OK
    if "合计" in row[0]: bg=C_F0
    for c,val in enumerate(row,1):
        bold=(c==1) or ("合计" in str(row[0]))
        fc=C_DARK
        if "🔴" in str(val): fc=C_RED; bold=True
        dat(ws5,r,c,val,bg=bg,bold=bold,color=fc,ah="center" if c>1 else "left")
    ws5.row_dimensions[r].height=18
cw(ws5,{1:22,2:14,3:14,4:16,5:16,6:12})

# ── Sheet 6: Pricing ───────────────────────────────────────────────────
ws6 = wb.create_sheet("Pricing Adjustments")
ws6.sheet_view.showGridLines = False
for c in range(1,9): ws6.cell(row=1,column=c).fill=fill(C_DARK)
ws6.merge_cells("A1:H1")
t6=ws6.cell(row=1,column=1,value="调价建议  |  PRICING ADJUSTMENT RECOMMENDATIONS")
t6.font=Font(bold=True,size=13,color="FFFFFF",name="Calibri"); t6.alignment=aln("center","center")
ws6.row_dimensions[1].height=28

sec(ws6,3,1,"Rate Sheet 调价对比  |  RATE SHEET ADJUSTMENT",merge=8)
hdrs8=["服务项目","当前单价","建议单价","调整幅度","年度增收估算","EBITA率提升","市场参考价","操作"]
for c,h in enumerate(hdrs8,1): hdr(ws6,4,c,h)

pricing=[
    ("入库操作 / Inbound Processing","8.16 RMB/行","8.80 RMB/行","+7.8%","+132K/年","+1.7%","7-10 RMB/行","Rate Sheet"),
    ("出库操作 / Outbound Processing","6.54 RMB/行","6.90 RMB/行","+5.5%","+252K/年","+3.2%","6-9 RMB/行","Rate Sheet"),
    ("中文贴标 / Chinese Labelling","0.15 RMB/标签","0.22 RMB/标签","+46.7%","+138K/年","+1.7%","0.20-0.30 RMB/标签","Rate Sheet"),
    ("DG Warranty Return","26.81 RMB/行","32.00 RMB/行","+19.4%","+4K/年","+0.05%","25-40 RMB/行","Rate Sheet"),
    ("**DG入库处理（新增）**","0（新增）","10-12 RMB/行","NEW","+132K~+158K","+1.7%~+2.0%","10-15 RMB/行","Rate Sheet 新增"),
    ("**DG出库处理（新增）**","0（新增）","12-15 RMB/行","NEW","+171K~+214K","+2.2%~+2.7%","12-18 RMB/行","Rate Sheet 新增"),
    ("综合调价合计","—","—","—","+829K~+878K","+10.4%~+11.0%","—","Rate Sheet 更新"),
]
for i,row in enumerate(pricing):
    r=5+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    if "综合" in row[0]: bg=C_OK
    elif "新增" in row[0]: bg=C_WO
    for c,val in enumerate(row,1):
        bold=(c==1) or ("综合" in str(row[0]))
        fc=C_DARK
        if "新增" in str(val): fc=C_RED; bold=True
        dat(ws6,r,c,val,bg=bg,bold=bold,color=fc,ah="center" if c in (2,3,4,5,6) else "left")
    ws6.row_dimensions[r].height=20

# Cost optimization section
sec(ws6,13,1,"成本优化建议  |  COST OPTIMIZATION",merge=8)
hdrs9=["优化项","方案","年度节省","可行性","说明"]
for c,h in enumerate(hdrs9,1): hdr(ws6,14,c,h)

opts=[
    ("减少1名 Supervisor","Parameter已标注，采纳建议","+313K/年","🟢 高","间接人工优化"),
    ("R-Truck租赁议价","3台 R-Truck 年租230,4K，折扣10%","+23K/年","🟡 中","需谈判能力"),
    ("安保系统报价核实","当前644K，3PL自建400-500K","一次性-150K","🟡 中","影响启动成本"),
    ("Packing Station共享","减少专用工位","一次性-30K","🟡 中","需运营评估"),
    ("薪资上调10%","降低流失率，减少 overtime","间接+100-150K","🟢 高","EBITA换质量"),
]
for i,row in enumerate(opts):
    r=15+i
    bg=C_F8 if i%2==0 else "FFFFFF"
    for c,val in enumerate(row,1):
        bold=(c==1)
        fc=C_DARK
        dat(ws6,r,c,val,bg=bg,bold=bold,color=fc,ah="center" if c in (3,4) else "left")
    ws6.row_dimensions[r].height=18
cw(ws6,{1:28,2:28,3:18,4:10,5:28,6:16,7:16,8:16})

wb.save(OUTPUT)
print(f"Saved: {OUTPUT}")
